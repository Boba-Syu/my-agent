"""
记账数据导出工具
支持将记账数据导出为 Excel（.xlsx）或 Markdown（.md）格式。
"""

from __future__ import annotations

import json
import logging
import os
from datetime import date
from pathlib import Path

from langchain_core.tools import tool

from app.db.sqlite_client import SQLiteClient

logger = logging.getLogger(__name__)

# 默认导出目录
_DEFAULT_EXPORT_DIR = Path("./data/exports")


def _ensure_export_dir(export_dir: str = "") -> Path:
    """确保导出目录存在并返回 Path 对象"""
    target = Path(export_dir) if export_dir else _DEFAULT_EXPORT_DIR
    target.mkdir(parents=True, exist_ok=True)
    return target


def _fetch_data(
    transaction_type: str | None,
    start_date: str | None,
    end_date: str | None,
    limit: int = 10000,
) -> list[dict]:
    """从数据库获取记账数据"""
    conditions = []
    params: dict = {"limit": limit}

    if transaction_type:
        conditions.append("transaction_type = :transaction_type")
        params["transaction_type"] = transaction_type
    if start_date:
        conditions.append("transaction_date >= :start_date")
        params["start_date"] = start_date
    if end_date:
        conditions.append("transaction_date <= :end_date")
        params["end_date"] = end_date

    where_clause = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    sql = (
        f"SELECT id, transaction_type, category, amount, note, transaction_date, created_at "
        f"FROM transactions {where_clause} "
        f"ORDER BY transaction_date DESC, id DESC LIMIT :limit"
    )

    db = SQLiteClient()
    return db.query(sql, params)


@tool
def export_to_excel(
    file_name: str = "",
    transaction_type: str = "",
    start_date: str = "",
    end_date: str = "",
    export_dir: str = "",
) -> str:
    """
    将记账数据导出为 Excel（.xlsx）文件。
    需要安装 openpyxl 库（已包含在项目依赖中）。

    Args:
        file_name:        导出文件名（不含扩展名），默认为 'accounting_YYYY-MM-DD'
        transaction_type: 过滤交易类型，'expense'（支出）/ 'income'（收入），为空时导出全部
        start_date:       起始日期过滤，格式 'YYYY-MM-DD'（可选）
        end_date:         结束日期过滤，格式 'YYYY-MM-DD'（可选）
        export_dir:       导出目录路径（可选，默认 ./data/exports/）

    Returns:
        导出文件的绝对路径，或错误说明
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return "❌ 缺少依赖：请先运行 uv add openpyxl"

    rows = _fetch_data(
        transaction_type=transaction_type or None,
        start_date=start_date or None,
        end_date=end_date or None,
    )

    if not rows:
        return "⚠️ 没有符合条件的记账数据可导出"

    # 构建文件路径
    today_str = date.today().strftime("%Y-%m-%d")
    name = file_name or f"accounting_{today_str}"
    export_path = _ensure_export_dir(export_dir) / f"{name}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "记账流水"

    # 表头
    headers = ["ID", "类型", "分类", "金额(¥)", "备注", "交易日期", "创建时间"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 类型中文映射
    type_cn = {"expense": "支出", "income": "收入"}

    # 数据行
    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row.get("id"))
        ws.cell(row=row_idx, column=2, value=type_cn.get(row.get("transaction_type", ""), row.get("transaction_type", "")))
        ws.cell(row=row_idx, column=3, value=row.get("category", ""))
        ws.cell(row=row_idx, column=4, value=row.get("amount", 0))
        ws.cell(row=row_idx, column=5, value=row.get("note", ""))
        ws.cell(row=row_idx, column=6, value=row.get("transaction_date", ""))
        ws.cell(row=row_idx, column=7, value=str(row.get("created_at", "")))

    # 调整列宽
    col_widths = [8, 8, 10, 12, 20, 14, 22]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 汇总行
    summary_row = len(rows) + 2
    ws.cell(row=summary_row, column=3, value="合计").font = Font(bold=True)
    ws.cell(row=summary_row, column=4, value=f"=SUM(D2:D{len(rows) + 1})").font = Font(bold=True)

    wb.save(export_path)
    abs_path = str(export_path.resolve())
    logger.info(f"Excel 已导出: {abs_path}，共 {len(rows)} 条记录")
    return f"✅ Excel 导出成功！\n文件路径: {abs_path}\n共导出 {len(rows)} 条记录"


@tool
def export_to_markdown(
    file_name: str = "",
    transaction_type: str = "",
    start_date: str = "",
    end_date: str = "",
    export_dir: str = "",
    include_summary: bool = True,
) -> str:
    """
    将记账数据导出为 Markdown（.md）文件，包含汇总统计表格。

    Args:
        file_name:        导出文件名（不含扩展名），默认为 'accounting_YYYY-MM-DD'
        transaction_type: 过滤交易类型，'expense'（支出）/ 'income'（收入），为空时导出全部
        start_date:       起始日期过滤，格式 'YYYY-MM-DD'（可选）
        end_date:         结束日期过滤，格式 'YYYY-MM-DD'（可选）
        export_dir:       导出目录路径（可选，默认 ./data/exports/）
        include_summary:  是否在文件末尾附加分类汇总（默认 True）

    Returns:
        导出文件的绝对路径，或错误说明
    """
    rows = _fetch_data(
        transaction_type=transaction_type or None,
        start_date=start_date or None,
        end_date=end_date or None,
    )

    if not rows:
        return "⚠️ 没有符合条件的记账数据可导出"

    today_str = date.today().strftime("%Y-%m-%d")
    name = file_name or f"accounting_{today_str}"
    export_path = _ensure_export_dir(export_dir) / f"{name}.md"

    type_cn = {"expense": "支出", "income": "收入"}
    lines: list[str] = []

    # 文档标题
    filter_desc = ""
    if start_date or end_date:
        filter_desc = f"（{start_date or '...'} ~ {end_date or '...'}）"
    if transaction_type:
        filter_desc += f"【{type_cn.get(transaction_type, transaction_type)}】"
    lines.append(f"# 记账明细{filter_desc}\n")
    lines.append(f"> 导出时间：{today_str}，共 {len(rows)} 条记录\n")
    lines.append("")

    # 明细表格
    lines.append("## 记账明细\n")
    lines.append("| ID | 类型 | 分类 | 金额(¥) | 备注 | 交易日期 |")
    lines.append("|---:|:----:|:----:|--------:|:-----|:--------:|")

    income_total = 0.0
    expense_total = 0.0

    for row in rows:
        t_type = type_cn.get(row.get("transaction_type", ""), row.get("transaction_type", ""))
        amount = row.get("amount", 0)
        if row.get("transaction_type") == "income":
            income_total += amount
        else:
            expense_total += amount
        lines.append(
            f"| {row.get('id')} | {t_type} | {row.get('category', '')} | "
            f"{amount:.2f} | {row.get('note', '') or '-'} | {row.get('transaction_date', '')} |"
        )

    lines.append("")

    # 收支汇总
    lines.append("## 收支汇总\n")
    lines.append("| 项目 | 金额(¥) |")
    lines.append("|:-----|--------:|")
    lines.append(f"| 💰 总收入 | {income_total:.2f} |")
    lines.append(f"| 💸 总支出 | {expense_total:.2f} |")
    net = income_total - expense_total
    net_str = f"+{net:.2f}" if net >= 0 else f"{net:.2f}"
    lines.append(f"| 📈 净额 | {net_str} |")
    lines.append("")

    # 分类汇总
    if include_summary:
        db = SQLiteClient()
        cat_rows = db.query(
            """
            SELECT transaction_type, category, SUM(amount) as total, COUNT(*) as count
            FROM transactions
            WHERE 1=1
              {type_filter}
              {date_filter_start}
              {date_filter_end}
            GROUP BY transaction_type, category
            ORDER BY transaction_type, total DESC
            """.format(
                type_filter=f"AND transaction_type = '{transaction_type}'" if transaction_type else "",
                date_filter_start=f"AND transaction_date >= '{start_date}'" if start_date else "",
                date_filter_end=f"AND transaction_date <= '{end_date}'" if end_date else "",
            )
        )

        if cat_rows:
            lines.append("## 分类汇总\n")
            lines.append("| 类型 | 分类 | 金额(¥) | 笔数 |")
            lines.append("|:----:|:----:|--------:|-----:|")
            for r in cat_rows:
                t_type = type_cn.get(r.get("transaction_type", ""), r.get("transaction_type", ""))
                lines.append(f"| {t_type} | {r.get('category', '')} | {r.get('total', 0):.2f} | {r.get('count', 0)} |")
            lines.append("")

    content = "\n".join(lines)
    export_path.write_text(content, encoding="utf-8")

    abs_path = str(export_path.resolve())
    logger.info(f"Markdown 已导出: {abs_path}，共 {len(rows)} 条记录")
    return f"✅ Markdown 导出成功！\n文件路径: {abs_path}\n共导出 {len(rows)} 条记录"
